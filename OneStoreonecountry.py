# import pandas as pd
# from pathlib import Path
# from openpyxl import load_workbook
# import os


# # Paths
# folder_paths = [
#     "C:/All Sales/Sales/NW",
#     "C:/All Sales/Sales/KC",
#     "C:/All Sales/Sales/SP",
#     "C:/All Sales/Sales/JM"
# ]
# file_path = "C:/All Sales/Sales - 2-12-2024/Daily Sales Report 02-12-24.xlsx"


# # Initialize a dictionary to store data by country
# data_by_country = {}

# # Check if the folder exists
# if not os.path.exists(folder_path):
#     print(f"Error: Folder '{folder_path}' does not exist.")
# else:
   
#     # List files in the folder
#     files = os.listdir(folder_path)
#     # print("Files in folder:", files)

#     # Process each file
#     for file_nm in files:
#         file_pt = os.path.join(folder_path, file_nm)
        
#         # Extract country name (assuming format: "JM COUNTRY.ext")
#         country_name = file_nm.split()[1].split(".")[0] if len(file_nm.split()) > 1 else "Unknown"

#         if file_nm.endswith(".csv"):
#             try:
#                 # Use pandas to read CSV files
#                 csv_data = pd.read_csv(file_pt)
#                 # print(f"Successfully loaded CSV file: {file_nm}")
#                 # Store CSV data as a DataFrame in the dictionary
#                 data_by_country[country_name] = csv_data
#             except Exception as e:
#                 print(f"Error loading CSV file {file_nm}: {e}")
        
#         else:
#             print(f"Unsupported file format: {file_nm}")

#     # Print the dictionary keys to verify
#     # print("Data by Country:", data_by_country.keys())

#     # Example: Access data for a specific country
#     # for country, data in data_by_country.items():
#     #     # print(f"Data for {country}:")
#     #     if isinstance(data, pd.DataFrame):  # Check if it's a DataFrame (CSV)
#     #         print(data.head())  # Display first few rows
#     #     else:  # It's a list (Excel)
#     #         for row in data[:5]:  # Display first 5 rows
#     #             print(row)


# # Define the country mapping dictionary
# country_mapping_list = {
#    "JM": {"AUS":"J M LIMITED (AUSTRALIA)",
#     "UK":"J M LIMITED (UK)",
#     "GER":"J M LIMITED (GERMANY)",
#     "SWE":"J M LIMITED (GERMANY)",
#     "BEL":"J M LIMITED (GERMANY)",
#     "NL":"J M LIMITED (GERMANY)",
#     "POL":"J M LIMITED (GERMANY)",
#     "SPA":"J M LIMITED (GERMANY)",
#     "FRA":"J M LIMITED (FRANCE)",
#     "IT": "J M LIMITED (ITALY)",},
#     "NW":{"AUS":"NORTH WEST (AUSTRALIA)",
#     "UK":"NORTH WEST (UK)",
#     "GER":"NORTH WEST (GERMANY)",
#     "SWE":"NORTH WEST (GERMANY)",
#     "BEL":"NORTH WEST (GERMANY)",
#     "NL":"NORTH WEST (GERMANY)",
#     "POL":"NORTH WEST (GERMANY)",
#     "SPA":"NORTH WEST (GERMANY)",
#     "FRA":"NORTH WEST (FRANCE)",
#     "IT": "NORTH WEST (ITALY)",},
#     "SP": {"AUS":"SPETRA (AUSTRALIA)",
#     "UK":"SPETRA (UK)",
#     "GER":"SPETRA (GERMANY)",
#     "SWE":"SPETRA (GERMANY)",
#     "BEL":"SPETRA (GERMANY)",
#     "NL":"SPETRA (GERMANY)",
#     "POL":"SPETRA (GERMANY)",
#     "SPA":"SPETRA (GERMANY)",
#     "FRA":"SPETRA (FRANCE)",
#     "IT": "SPETRA (ITALY)",},
#     "KC":{"AUS":"KC STORE (AUSTRALIA)",
#     "UK":"KC STORE (UK)",
#     "GER":"KC STORE (GERMANY)",
#     "SWE":"KC STORE (GERMANY)",
#     "BEL":"KC STORE (GERMANY)",
#     "NL":"KC STORE (GERMANY)",
#     "POL":"KC STORE (GERMANY)",
#     "SPA":"KC STORE (GERMANY)",
#     "FRA":"KC STORE (FRANCE)",
#     "IT": "KC STORE (ITALY)"}
# }
# country_mapping=country_mapping_list[folder_name]
# print(country_mapping)
# # Extract the folder name
# folder_name = Path(folder_path).name
# # print("Folder Name:", folder_name)

# # Define the mapping dictionary
# sheet_name_mapping = {
#     "KC": "KC Product Sales",
#     "NW": "North West Product Sales",
#     "JM": "J M LIMITED",
#     "SP": "Spetra Product Sales"
# }

# # Look up the sheet name using the folder name
# matched_sheet_name = sheet_name_mapping.get(folder_name)
# if matched_sheet_name:
#     # Load the Excel sheet into pandas with engine specified
#     data = pd.read_excel(file_path, sheet_name=matched_sheet_name, engine='openpyxl')

#      # Define country sections and start/end markers
#     for country_code, country_data in data_by_country.items():
#         print(f"..........................Processing data for country: {country_data}")

#         # Check if the country is in the country_mapping and proceed
#         section = country_mapping.get(country_code)
#         if section:
#             print(f"Processing section: {section}")

#             # Strip any leading or trailing spaces from the column with country names
#             data.iloc[:, 1] = data.iloc[:, 1].str.strip()

#             # Find the row index where the country section starts
#             section_indices = data[data.iloc[:, 1] == section].index

#             if len(section_indices) == 0:
#                 print(f"Section '{section}' not found in the data.")
#                 continue
#             else:
#                 start_row = section_indices[0] + 1
#                 print(f"Start row: {start_row}")

#          # Iterate over rows in the section to find the first blank row
#             end_row = start_row  # Initialize end_row to start_row
#             for idx in range(start_row, len(data)):
#                 if pd.isna(data.iloc[idx, 0]):  # If the first column is empty (blank row)
#                     end_row = idx
#                     print(f"Blank row found at index {idx}, setting end_row to {end_row}")
#                     break

#             print(f"End row: {end_row}")
   
#         for i,row in country_data.iterrows():
#             # Access 'Child ASIN' and 'unit ordered' columns
#             target_code=row['(Child) ASIN']
#             unit_ordered = row['Units ordered']
#             # Extract rows for the target section
#             section_data = data.iloc[start_row:end_row].reset_index(drop=True)
#            # print(section_data.iloc[:, 1].str.strip())
#             # Search for the target code
#             found_row = section_data[section_data.iloc[:, 0].str.strip() == target_code]
#             if not found_row.empty:
#                  # Get the relative row index in the section_data
#                 relative_row_index = found_row.index[0]

#                 # Calculate the absolute row index in the original data
#                 absolute_row_index = (start_row+2) + relative_row_index
#                 workbook = load_workbook(file_path)
#                 sheet = workbook[matched_sheet_name]  # Replace with the correct sheet name if needed

#                 sheet.cell(row=absolute_row_index, column=6).value = unit_ordered

#                 workbook.save(file_path)

#                 print("Excel file updated successfully.")
#                 print(f"Found code '{target_code}' at row {absolute_row_index} in section '{section}'.")
#                 print(f"Code '{target_code}' found in section '{section}':")
#                 # print(found_row)
#             else:
#                 print(f"Code '{target_code}' not found in section '{section}'.")
            
            
           

























#     # # Extract country-specific data
#     # for i, section in enumerate(country_sections):
#     #     print(f"Processing section: {section}")

#     #     # Strip any leading or trailing spaces from the column with country names
#     #     data.iloc[:, 1] = data.iloc[:, 1].str.strip()

#     #     # Find the row index where the country section starts
#     #     section_indices = data[data.iloc[:, 1] == section].index

#     #     if len(section_indices) == 0:
#     #         print(f"Section '{section}' not found in the data.")
#     #         continue
#     #     else:
#     #         start_row = section_indices[0] + 1
#     #         print(f"Start row: {start_row}")
#     #         # Determine the end row (next section or end of file)
#     #         if i < len(country_sections) - 1:
#     #             next_section_indices = data[data.iloc[:, 1] == country_sections[i + 1]].index
#     #             end_row = next_section_indices[0] if len(next_section_indices) > 0 else len(data)
#     #         else:
#     #             end_row = len(data)

#     #         print(f"End row: {end_row}")
#     #         print(matched_sheet_name)
#     #         workbook = load_workbook(file_path)
#     #         sheet = workbook[matched_sheet_name]  # Replace with the correct sheet name if needed

#     #         #code searching .....................................................................
#     #         target_code="B01CMJZTNO"
#     #         # Extract rows for the target section
#     #         section_data = data.iloc[start_row:end_row].reset_index(drop=True)
#     #        # print(section_data.iloc[:, 1].str.strip())
#     #         # Search for the target code
#     #         found_row = section_data[section_data.iloc[:, 0].str.strip() == target_code]

#     #         if not found_row.empty:
#     #              # Get the relative row index in the section_data
#     #             relative_row_index = found_row.index[0]

#     #             # Calculate the absolute row index in the original data
#     #             absolute_row_index = (start_row+2) + relative_row_index
#     #             sheet.cell(row=absolute_row_index, column=6).value = '100'

#     #             workbook.save(file_path)

#     #             print("Excel file updated successfully.")
#     #             print(f"Found code '{target_code}' at row {absolute_row_index} in section '{section}'.")
#     #             print(f"Code '{target_code}' found in section '{section}':")
#     #             # print(found_row)
#     #         else:
#     #             print(f"Code '{target_code}' not found in section '{section}'.")

# else:
#     print("No matching sheet name found.")








import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os

# Define multiple folder paths
folder_paths = [
    "C:/All Sales/Sales/NW",
    "C:/All Sales/Sales/KC",
    "C:/All Sales/Sales/SP",
    "C:/All Sales/Sales/JM"
]

# Define the base file path (assuming this file remains the same for all folders)
file_path = "C:/All Sales/Sales/Daily Sales Report.xlsx"

# Mapping for country-specific sections based on folder names
country_mapping_list = {
    "JM": {
        "AUS": "J M LIMITED (AUSTRALIA)",
        "UK": "J M LIMITED (UK)",
        "GER": "J M LIMITED (GERMANY)",
        "SWE": "J M LIMITED (GERMANY)",
        "BEL": "J M LIMITED (GERMANY)",
        "NL": "J M LIMITED (GERMANY)",
        "POL": "J M LIMITED (GERMANY)",
        "SPA": "J M LIMITED (GERMANY)",
        "FRA": "J M LIMITED (FRANCE)",
        "IT": "J M LIMITED (ITALY)"
    },
    "NW": {
        "AUS": "NORTH WEST (AUSTRALIA)",
        "UK": "NORTH WEST (UK)",
        "GER": "NORTH WEST (GERMANY)",
        "SWE": "NORTH WEST (GERMANY)",
        "BEL": "NORTH WEST (GERMANY)",
        "NL": "NORTH WEST (GERMANY)",
        "POL": "NORTH WEST (GERMANY)",
        "SPA": "NORTH WEST (GERMANY)",
        "FRA": "NORTH WEST (FRANCE)",
        "IT": "NORTH WEST (ITALY)"
    },
    "SP": {
        "AUS": "SPETRA (AUSTRALIA)",
        "UK": "SPETRA (UK)",
        "GER": "SPETRA (GERMANY)",
        "SWE": "SPETRA (GERMANY)",
        "BEL": "SPETRA (GERMANY)",
        "NL": "SPETRA (GERMANY)",
        "POL": "SPETRA (GERMANY)",
        "SPA": "SPETRA (GERMANY)",
        "FRA": "SPETRA (FRANCE)",
        "IT": "SPETRA (ITALY)"
    },
    "KC": {
        "AUS": "KC STORE (AUSTRALIA)",
        "UK": "KC STORE (UK)",
        "GER": "KC STORE (GERMANY)",
        "SWE": "KC STORE (GERMANY)",
        "BEL": "KC STORE (GERMANY)",
        "NL": "KC STORE (GERMANY)",
        "POL": "KC STORE (GERMANY)",
        "SPA": "KC STORE (GERMANY)",
        "FRA": "KC STORE (FRANCE)",
        "IT": "KC STORE (ITALY)"
    }
}

# Mapping for sheet names based on folder names
sheet_name_mapping = {
    "KC": "KC Product Sales",
    "NW": "North West Product Sales",
    "JM": "J M LIMITED",
    "SP": "Spetra Product Sales"
}

# Process each folder
for folder_path in folder_paths:
    print(f"Processing folder: {folder_path}")
    folder_name = Path(folder_path).name  # Extract the folder name
    country_mapping = country_mapping_list.get(folder_name)  # Get the country mapping for this folder
    matched_sheet_name = sheet_name_mapping.get(folder_name)  # Get the corresponding sheet name

    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        continue

    if not matched_sheet_name:
        print(f"Error: No matching sheet name found for folder: {folder_name}")
        continue

    # Load Excel sheet
    try:
        data = pd.read_excel(file_path, sheet_name=matched_sheet_name, engine='openpyxl')
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        continue

    # Read CSV files in the folder
    data_by_country = {}
    files = os.listdir(folder_path)
    for file_nm in files:
        file_pt = os.path.join(folder_path, file_nm)

        if file_nm.endswith(".csv"):
            try:
                country_name = file_nm.split()[1].split(".")[0] if len(file_nm.split()) > 1 else "Unknown"
                csv_data = pd.read_csv(file_pt)
                data_by_country[country_name] = csv_data
            except Exception as e:
                print(f"Error loading CSV file {file_nm}: {e}")

    # Update Excel with data from CSVs
    for country_code, country_data in data_by_country.items():
        section = country_mapping.get(country_code)
        if not section:
            print(f"Warning: No section mapping found for country code: {country_code}")
            continue

        # Find section in the Excel sheet
        data.iloc[:, 1] = data.iloc[:, 1].str.strip()  # Clean up whitespace
        section_indices = data[data.iloc[:, 1] == section].index

        if len(section_indices) == 0:
            print(f"Warning: Section '{section}' not found in the sheet '{matched_sheet_name}'.")
            continue

        start_row = section_indices[0] + 1
        end_row = len(data)
        for idx in range(start_row, len(data)):
            if pd.isna(data.iloc[idx, 0]):  # Look for the first blank row
                end_row = idx
                break

        # Update cells in the Excel file
        section_data = data.iloc[start_row:end_row].reset_index(drop=True)
        for _, row in country_data.iterrows():
            target_code = row['(Child) ASIN']
            unit_ordered = row['Units ordered']

            found_row = section_data[section_data.iloc[:, 0].str.strip() == target_code]
            if not found_row.empty:
                relative_row_index = found_row.index[0]
                absolute_row_index = start_row + relative_row_index + 2

                try:
                    workbook = load_workbook(file_path)
                    sheet = workbook[matched_sheet_name]
                    cell_value = sheet.cell(row=absolute_row_index, column=6).value
                    # Check if the cell has a value
                    if cell_value is not None and cell_value != "":
                        sheet.cell(row=absolute_row_index, column=6).value += unit_ordered
                    else:
                       sheet.cell(row=absolute_row_index, column=6).value = unit_ordered
                    
                    workbook.save(file_path)
                except Exception as e:
                    print(f"Error updating Excel file: {e}")
                    continue

                #print(f"Updated code '{target_code}' in section '{section}' at row {absolute_row_index} and value {unit_ordered}.")
            else:
                print(f"Code '{target_code}' not found in section '{section}'.")
